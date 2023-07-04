from openpyxl import styles
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from selenium.webdriver.common.by import By
import re
import time

def logOrders(user, driver, url):
    
    driver.get(url) # open orders history
    get_buttons = driver.find_elements(By.CSS_SELECTOR, 'button[id^="OrderDetails"]')
    
    if len(get_buttons) == 0:
        return False
    
    for trans in get_buttons:
        trans.click()
        time.sleep(1)

    time.sleep(2)
    orders_log = driver.find_element(By.TAG_NAME, 'tbody')

    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Date'
    ws['B1'] = 'OrderId'
    ws['C1'] = 'CardNo./Pin'
    ws['D1'] = 'Amount'
    ws['E1'] = 'Platform'
    ws['F1'] = 'Expiration'

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 12

    for rew in re.split('Redeemed|Sent', orders_log.text):
        transactional_details = []
        rew_split = rew.split("\n")
        for j, k in enumerate(rew_split):
            info = k.split()
            if len(info) >= 1 and re.match(r'\d\d?\/\d\d?\/\d{4}', info[0]):
                month, day, year = info[0].split('/')
                transactional_details.append(f'{day.zfill(2)}/{month.zfill(2)}/{year}')
            if 'Amazon' in info or 'Flipkart' in info:
                transactional_details.append(info[0])
                transactional_details.append(info[1])
            if 'Order' in info and len(info)>2:
                transactional_details.append(info[2])
            if 'Code:' in info or 'Number:'in info:
                if 'Code:' in info:
                    transactional_details.append(rew_split[rew_split.index(k)+1])
                elif 'Number:'in info:
                    transactional_details.append(f"{rew_split[rew_split.index(k)+1]} - {rew_split[rew_split.index(k)-1]}")
            if 'Expiration' in info:
                transactional_details.append(rew_split[rew_split.index(k)+1].split('T')[0])

        if len(transactional_details) > 2:
            if len(transactional_details) != 6:
                transactional_details += ([0] * (6 - len(transactional_details)))

            transactional_details[1], transactional_details[3] = transactional_details[3],  transactional_details[1]
            transactional_details[2], transactional_details[4] = transactional_details[4],  transactional_details[2]
            transactional_details[3] = int(transactional_details[3][1:])
            
            ws.append(transactional_details)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = styles.Alignment(horizontal='center', vertical='center')

    tab = Table(displayName="orders_history", ref=f"A1:F{ws.max_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    ws.add_table(tab)
    wb.save(f"redeem/{user.split('@')[0]}.xlsx")
    return True